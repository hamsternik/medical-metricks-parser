__author__ = 'khomitsevich'

import sys
import math
import pprint
import openpyxl

# TODO: Move out all logic of parsing into standalone class/module
import re

from metricks.os_helper import OSHelper
from metricks.dir_scanner import DirectoryScanner
from metricks.result_worksheet_data_processor import ResultWorkSheetDataProcessor

import metricks.operations.metricks_operations as metricks_operations

### Path to files, will be removed after the release
HARDCODED_DIR_PATH = '/Users/hamsternik/Downloads/metricks'

def main(path: str):
    dir_scanner = DirectoryScanner(path)
    files_dict = dir_scanner.filesDictToProcess
    # pprint.pprint(files_dict) ### DEBUG

    metricks_dict = metricks_operations.get_metricks_dict(files_dict)
    # pprint.pprint(metrics_dict) ### DEBUG

    """ Get sorted list of folder titles contains metrics data. """
    metrics_dict_keys_sorted_list = metricks_operations.get_metricks_dict_keys_sorted_list(metricks_dict)

    """ Starting work with .excel file. """
    workbook = openpyxl.Workbook()
    
    DATA_SHEET_TITLE = "Data_Sheet"
    active_worksheet = workbook.create_sheet(title=DATA_SHEET_TITLE, index=1)

    PARAMS_EMPTY_SPACE_OFFSET = 1 # Empty cell to left a space for parameter titles (fill out below).
    INSTANCE_NUMBER_AS_INDEX_OFFSET = 1 # Empty cell to left a space for each metrics instance number (filled out below).
    DIR_TITLE_START_INDEX = (PARAMS_EMPTY_SPACE_OFFSET + INSTANCE_NUMBER_AS_INDEX_OFFSET) + 1 # Start index position for the first metrics folder (title).

    """ Fill out .excel file with the names of all scanned directories by them titles, ex. `TCGA-08-0392` """
    for i in range(DIR_TITLE_START_INDEX, PARAMS_EMPTY_SPACE_OFFSET + (INSTANCE_NUMBER_AS_INDEX_OFFSET + 1) * len(metrics_dict_keys_sorted_list) + 1):
        """ 
            Fill .excel file with folder titles.
            start: first folder title should be after at 3rd position, after the parameter titles and a gap for metrics number
            end: params_shift + len(metrics_dict.keys()) * 2 (folder title + cell for metrics number) + 1 (take into account the last element)
        """
        if i == 1 or i % 2 == 0:
            pass
        if i % 2 == 1:
            active_worksheet.cell(1, i, metrics_dict_keys_sorted_list[(i // 2) - 1])
    
    """ Find the biggest number of metrics instances in one read folder. """
    max_metrics_number_overall = metricks_operations.get_max_metricks_number_overall(metricks_dict)

    """ Define list of metrics parameter titles (dice, spec, accur, etc.). """
    METRICKS_PARAM_TITLES = ["DICE", "SPEC", "SENS", "ACCUR", "Time (w init)", "Time (w/o init)", "Otsu Treshold", "Average In", "Average Out"]

    EMPTY_CELLS_LEADING_START_INDEX = 2 # How many cells need to left empty.
    PARAMETER_TRAILING_OFFSET = 1  # Parameters enumeration starts from the empty cell.
    PARAMETER_LEADING_OFFSET = 1 # Parameters enumaration ends with the empty cell.

    """ Fill out the first column with metrics parameter titles, ex. `DICE`, `ACCUR`. """
    for i in range(max_metrics_number_overall):
        start_index = (PARAMETER_TRAILING_OFFSET + len(METRICKS_PARAM_TITLES) + PARAMETER_LEADING_OFFSET) * i
        end_index = len(METRICKS_PARAM_TITLES) + PARAMETER_LEADING_OFFSET
        for j in range(end_index + 1):
            row_index = EMPTY_CELLS_LEADING_START_INDEX + start_index + j
            cell_value = None if j == 0 or j == end_index else METRICKS_PARAM_TITLES[j-1]
            active_worksheet.cell(row_index, 1, value=cell_value)
    
    """ Get and sort all keys (absolute filepathes) for each of scanned directories. """
    metrics_dict_keys_full_sorted_list = [full_path for full_path in metricks_dict.keys()]
    metrics_dict_keys_full_sorted_list.sort()

    """ Fill out metrics title as `4 - metrics` (only title indexes) for each scanned directory. """
    for k in range(len(metrics_dict_keys_full_sorted_list)):
        metrics_list = metricks_dict[metrics_dict_keys_full_sorted_list[k]]
        for i in range(len(metrics_list)):
            start_index = (PARAMETER_TRAILING_OFFSET + len(METRICKS_PARAM_TITLES) + PARAMETER_LEADING_OFFSET) * i
            end_index = len(METRICKS_PARAM_TITLES) + PARAMETER_LEADING_OFFSET
            
            """ Write for each metrics instance according index (parsed from name). """
            for j in range(end_index):
                row_index = EMPTY_CELLS_LEADING_START_INDEX + start_index + j
                cur_metrics_name = metrics_list[i].name
                cell_value = int(re.search(r'\d+', cur_metrics_name).group()) if j == 0 else None 
                active_worksheet.cell(row_index, 2 + 2*k, value=cell_value)
            
            """ Write metrics parameter values. """
            for j in range(end_index):
                row_index = EMPTY_CELLS_LEADING_START_INDEX + start_index + j
                cur_metrics_output_params_list = [metrics_list[i].dice, metrics_list[i].spec, metrics_list[i].sens, 
                                                metrics_list[i].accu, metrics_list[i].time_of_work_with_init, 
                                                metrics_list[i].time_of_work_wihout_init, metrics_list[i].otsu_treshold, 
                                                metrics_list[i].average_in, metrics_list[i].average_out]
                cell_value = "ALG Results:" if j == 0 else cur_metrics_output_params_list[j-1]
                active_worksheet.cell(row_index, 2 + (2*k + 1), value=cell_value)

    result_worksheet = workbook.create_sheet(title="Result_Sheet", index=2)

    METRICKS_DIR_TITLES_EXCEL_INDEX_START = 3
    METRICKS_DIR_TITLES_EXCEL_INDEX_END = METRICKS_DIR_TITLES_EXCEL_INDEX_START + len(metrics_dict_keys_sorted_list)

    for i in range(METRICKS_DIR_TITLES_EXCEL_INDEX_START, METRICKS_DIR_TITLES_EXCEL_INDEX_END):
        """ Form headers (metricks directory titles) for next generated data. """
        result_worksheet.cell(1, i + 1, metrics_dict_keys_sorted_list[i - METRICKS_DIR_TITLES_EXCEL_INDEX_START])

    """ Define list of metrics parameter titles (dice, spec, accur, etc.). """
    RESULT_PARAM_TITLES = ["DICE", "SPEC", "SENS", "ACCUR", "Time (w init)"]
    
    RESULT_FUNC_TITLE_IND_OFFSET = 1
    RESULT_FUNC_END_IND = RESULT_FUNC_TITLE_IND_OFFSET * 2 + len(RESULT_PARAM_TITLES) * 2

    for i in range(RESULT_FUNC_END_IND):
        if i == RESULT_FUNC_TITLE_IND_OFFSET:
            result_worksheet.cell(i+1, 1, "AV")
        elif i == RESULT_FUNC_TITLE_IND_OFFSET * 2 + len(RESULT_PARAM_TITLES):
            result_worksheet.cell(i+1, 1, "STD")
    
    RESULT_PARAM_COLUMN_LEADING_IND = 3

    for i in range(RESULT_FUNC_END_IND):
        if i >= RESULT_FUNC_TITLE_IND_OFFSET and i < RESULT_FUNC_TITLE_IND_OFFSET + len(RESULT_PARAM_TITLES):
            list_ind = i - RESULT_FUNC_TITLE_IND_OFFSET
            result_worksheet.cell(i+1, RESULT_PARAM_COLUMN_LEADING_IND, RESULT_PARAM_TITLES[list_ind])
        if i >= RESULT_FUNC_TITLE_IND_OFFSET * 2 + len(RESULT_PARAM_TITLES) and i < RESULT_FUNC_TITLE_IND_OFFSET * 2 + len(RESULT_PARAM_TITLES) + len(RESULT_PARAM_TITLES):
            list_ind = i - (RESULT_FUNC_TITLE_IND_OFFSET * 2 + len(RESULT_PARAM_TITLES))
            result_worksheet.cell(i+1, RESULT_PARAM_COLUMN_LEADING_IND, RESULT_PARAM_TITLES[list_ind])

    dice_indexes_dict = {}
    spec_indexes_dict = {}
    sens_indexes_dict = {}
    accur_indexes_dict = {}
    time_indexes_dict = {}
    for k in range(len(metrics_dict_keys_full_sorted_list)):
        dice_indexes_list = []
        spec_indexes_list = []
        sens_indexes_list = []
        accur_indexes_list = []
        time_indexes_list = []
        metrics_list = metricks_dict[metrics_dict_keys_full_sorted_list[k]]
        for i in range(len(metrics_list)):
            start_index = (PARAMETER_TRAILING_OFFSET 
                + len(METRICKS_PARAM_TITLES) 
                + PARAMETER_LEADING_OFFSET) * i
            end_index = len(METRICKS_PARAM_TITLES) + PARAMETER_LEADING_OFFSET
            for j in range(end_index):
                """ Write metrics parameter values. """
                row_index = EMPTY_CELLS_LEADING_START_INDEX + start_index + j
                col_index = 2 + 2*k
                cur_metrics_output_params_list = [metrics_list[i].dice, metrics_list[i].spec, metrics_list[i].sens, 
                                                metrics_list[i].accu, metrics_list[i].time_of_work_with_init, 
                                                metrics_list[i].time_of_work_wihout_init, metrics_list[i].otsu_treshold, 
                                                metrics_list[i].average_in, metrics_list[i].average_out]
                if j == 1:
                    dice_indexes_list.append( (row_index, col_index) )
                elif j == 2:
                    spec_indexes_list.append( (row_index, col_index) )
                elif j == 3:
                    sens_indexes_list.append( (row_index, col_index) )
                elif j == 4:
                    accur_indexes_list.append( (row_index, col_index) )
                elif j == 5:
                    time_indexes_list.append( (row_index, col_index) )
        
        dice_indexes_dict[metrics_dict_keys_full_sorted_list[k]] = dice_indexes_list
        spec_indexes_dict[metrics_dict_keys_full_sorted_list[k]] = spec_indexes_list
        sens_indexes_dict[metrics_dict_keys_full_sorted_list[k]] = sens_indexes_list
        accur_indexes_dict[metrics_dict_keys_full_sorted_list[k]] = accur_indexes_list
        time_indexes_dict[metrics_dict_keys_full_sorted_list[k]] = time_indexes_list
    
    # pprint.pprint(dice_indexes_dict) ##3 DEBUG

    statistical_result_sheet = ResultWorkSheetDataProcessor(
        result_worksheet, 
        f"{DATA_SHEET_TITLE}!", 
        metrics_dict_keys_full_sorted_list
    )
    statistical_result_sheet.calculate_average_and_std_of_dice(dice_indexes_dict)
    statistical_result_sheet.calculate_average_and_std_of_spec(spec_indexes_dict)
    statistical_result_sheet.calculate_average_and_std_of_sens(sens_indexes_dict)
    statistical_result_sheet.calculate_average_and_std_of_accur(accur_indexes_dict)
    statistical_result_sheet.calculate_average_and_std_of_time(time_indexes_dict)

    workbook.save("../result.xlsx")
# def main

if __name__ == '__main__':
    input_args = sys.argv
    if len(input_args) == 2:
        OSHelper.validateInputArgumentsAmount(input_args)
        OSHelper.validateInputArgumentAsDirectory(input_args[1])
        main(input_args[1])
    else:
        # main(OSHelper.getCurrentDirectory)
        main(HARDCODED_DIR_PATH)
    print("Excel has been created successfully!")
# FINISH